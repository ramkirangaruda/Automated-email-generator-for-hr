#!/usr/bin/env python3
"""
Employee data store for iFocus Birthday Automation
----------------------------------------------------
Single source of truth for employee records (name, email, dob), backed by a
local SQLite database. Used by both the HR portal (portal_app.py) and the
birthday email sender (send_birthday_emails.py), so edits made through the
portal are picked up by the next scheduled email run automatically.

Also handles importing employee rows from an uploaded .xlsx (bulk upload from
HR) and exporting the current table back out to .xlsx (backup/download).
"""

import os
import re
import sqlite3
from contextlib import contextmanager
from datetime import date, datetime

import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.environ.get("EMPLOYEES_DB", os.path.join(BASE_DIR, "employees.db"))

# Legacy xlsx files this project used before the DB existed - if the DB is
# brand new and empty, we auto-import from whichever of these is found so
# existing data isn't lost.
_LEGACY_XLSX_CANDIDATES = [
    os.path.join(BASE_DIR, "employees.local.xlsx"),
    os.path.join(BASE_DIR, "employees.xlsx"),
]

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


class ValidationError(Exception):
    pass


@contextmanager
def _connect():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def init_db():
    """Create the employees table if missing, and auto-import legacy xlsx data
    the first time the DB is created (so nothing is lost in the transition)."""
    is_new = not os.path.exists(DB_PATH)

    with _connect() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS employees (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                email TEXT NOT NULL UNIQUE COLLATE NOCASE,
                dob TEXT NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )

    if is_new:
        for path in _LEGACY_XLSX_CANDIDATES:
            if os.path.exists(path):
                try:
                    import_from_xlsx(path)
                except Exception:
                    pass
                break


def _validate(name, email, dob_str):
    name = (name or "").strip()
    email = (email or "").strip()
    dob_str = (dob_str or "").strip()

    if not name:
        raise ValidationError("Name is required.")
    if not email or not EMAIL_RE.match(email):
        raise ValidationError(f"'{email}' is not a valid email address.")
    try:
        dob = date.fromisoformat(dob_str)
    except ValueError:
        raise ValidationError(f"'{dob_str}' is not a valid date (expected YYYY-MM-DD).")

    return name, email, dob.isoformat()


def list_employees():
    with _connect() as conn:
        rows = conn.execute(
            "SELECT * FROM employees ORDER BY name COLLATE NOCASE"
        ).fetchall()
        return [dict(r) for r in rows]


def get_employee(emp_id):
    with _connect() as conn:
        row = conn.execute("SELECT * FROM employees WHERE id = ?", (emp_id,)).fetchone()
        return dict(row) if row else None


def add_employee(name, email, dob_str):
    name, email, dob_iso = _validate(name, email, dob_str)
    now = datetime.now().isoformat(timespec="seconds")
    with _connect() as conn:
        try:
            conn.execute(
                "INSERT INTO employees (name, email, dob, created_at, updated_at) "
                "VALUES (?, ?, ?, ?, ?)",
                (name, email, dob_iso, now, now),
            )
        except sqlite3.IntegrityError:
            raise ValidationError(f"An employee with email '{email}' already exists.")


def update_employee(emp_id, name, email, dob_str):
    name, email, dob_iso = _validate(name, email, dob_str)
    now = datetime.now().isoformat(timespec="seconds")
    with _connect() as conn:
        try:
            cur = conn.execute(
                "UPDATE employees SET name = ?, email = ?, dob = ?, updated_at = ? "
                "WHERE id = ?",
                (name, email, dob_iso, now, emp_id),
            )
        except sqlite3.IntegrityError:
            raise ValidationError(f"An employee with email '{email}' already exists.")
        if cur.rowcount == 0:
            raise ValidationError("Employee not found.")


def delete_employee(emp_id):
    with _connect() as conn:
        cur = conn.execute("DELETE FROM employees WHERE id = ?", (emp_id,))
        if cur.rowcount == 0:
            raise ValidationError("Employee not found.")


def get_todays_birthdays(today):
    """Return list of (name, email) for people whose month/day matches today."""
    with _connect() as conn:
        rows = conn.execute("SELECT name, email, dob FROM employees").fetchall()

    matches = []
    for row in rows:
        try:
            dob = date.fromisoformat(row["dob"])
        except ValueError:
            continue
        if dob.month == today.month and dob.day == today.day:
            matches.append((row["name"], row["email"]))
    return matches


def _read_xlsx_rows(xlsx_path):
    """Yield (name, email, dob_str) tuples from an uploaded/legacy xlsx file.
    Expects a header row with 'name', 'email', 'dob' columns (any order,
    case-insensitive). Rows with missing/placeholder emails or unparseable
    dates are skipped and reported.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    try:
        header = next(rows)
    except StopIteration:
        wb.close()
        return

    header = [str(h).strip().lower() if h is not None else "" for h in header]
    try:
        name_idx = header.index("name")
        email_idx = header.index("email")
        dob_idx = header.index("dob")
    except ValueError:
        wb.close()
        raise ValidationError(
            f"Excel file must have 'name', 'email', 'dob' columns; found: {header}"
        )

    for row_num, row in enumerate(rows, start=2):
        if row is None or all(cell is None for cell in row):
            continue

        name = str(row[name_idx]).strip() if row[name_idx] is not None else ""
        email = str(row[email_idx]).strip() if row[email_idx] is not None else ""
        dob_raw = row[dob_idx]

        dob_str = None
        if isinstance(dob_raw, (date, datetime)):
            dob_str = dob_raw.date().isoformat() if isinstance(dob_raw, datetime) else dob_raw.isoformat()
        elif dob_raw is not None:
            try:
                dob_str = date.fromisoformat(str(dob_raw).strip()).isoformat()
            except ValueError:
                dob_str = None

        yield row_num, name, email, dob_str

    wb.close()


def import_from_xlsx(xlsx_path):
    """Bulk import/update from an uploaded Excel file. Matches existing
    employees by email (case-insensitive) and updates them; new emails are
    inserted. Returns a summary dict with counts and per-row skip reasons.
    """
    added = 0
    updated = 0
    skipped = []

    for row_num, name, email, dob_str in _read_xlsx_rows(xlsx_path):
        if not name:
            skipped.append((row_num, "missing name"))
            continue
        if not email or "TODO" in email.upper() or not EMAIL_RE.match(email):
            skipped.append((row_num, f"invalid/placeholder email ({email!r})"))
            continue
        if not dob_str:
            skipped.append((row_num, "unparseable date of birth"))
            continue

        existing = None
        with _connect() as conn:
            row = conn.execute(
                "SELECT id FROM employees WHERE email = ? COLLATE NOCASE", (email,)
            ).fetchone()
            existing = row["id"] if row else None

        if existing:
            update_employee(existing, name, email, dob_str)
            updated += 1
        else:
            add_employee(name, email, dob_str)
            added += 1

    return {"added": added, "updated": updated, "skipped": skipped}


def export_to_xlsx(xlsx_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "employees"
    ws.append(["name", "email", "dob"])
    for emp in list_employees():
        ws.append([emp["name"], emp["email"], emp["dob"]])
    wb.save(xlsx_path)
    return xlsx_path
